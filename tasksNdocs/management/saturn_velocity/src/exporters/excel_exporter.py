"""
Excel / JSON exporter — serialise calculation results and snapshots for download.
"""
import io
from datetime import date
from typing import List

from ..domain.models import CalculationResult, CalculationSnapshot, Sprint, Scenario

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ── Styling helpers ──────────────────────────────────────────────────────────

_HDR_FONT = None
_HDR_FILL = None
_WARN_FILL = None

def _styles():
    global _HDR_FONT, _HDR_FILL, _WARN_FILL
    if _HDR_FONT is None:
        _HDR_FONT = Font(bold=True, color="FFFFFF")
        _HDR_FILL = PatternFill("solid", fgColor="1F4E79")
        _WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
    return _HDR_FONT, _HDR_FILL, _WARN_FILL


def _hrow(ws, row: int, values: list):
    hf, hfill, _ = _styles()
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")


def _autowidth(ws, extra: int = 4, max_width: int = 55):
    for col in ws.columns:
        best = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(best + extra, max_width)


# ── XLSX export ──────────────────────────────────────────────────────────────

def result_to_xlsx(sprint: Sprint, scenario: Scenario, result: CalculationResult) -> bytes:
    """Export a CalculationResult to an XLSX workbook (bytes)."""
    if not _HAS_OPENPYXL:
        return b""

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    bf = Font(bold=True)

    meta = [
        ("Sprint", sprint.name),
        ("Scenario", scenario.name),
        ("Start Date", str(sprint.start_date)),
        ("End Date", str(sprint.end_date)),
        ("Development End", str(sprint.development_end_date)),
        ("Public Holidays", sprint.public_holidays),
        ("Buffer", sprint.buffer),
        ("Backup (days)", sprint.backup),
        ("Fixed Day Deduction", sprint.fixed_day_deduction),
        ("Rule Version", result.rule_version),
        ("As-of Date", str(result.as_of_date)),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws.cell(r, 1, k).font = bf
        ws.cell(r, 2, v)

    r = len(meta) + 2
    _hrow(ws, r, ["KPI", "Value"])
    kpis = [
        ("Dev Days  [BR-01]", result.dev_days),
        ("Remaining Dev Days  [BR-02]", result.remaining_dev_days),
        ("Buffer Days  [BR-10]", result.buffer_days),
        ("Full Dev V  [BR-07]", result.full_dev_v),
        ("Dev V  [BR-07]", result.dev_v),
        ("Full QC V  [BR-07]", result.full_qc_v),
        ("QC V  [BR-07]", result.qc_v),
        ("Team Velocity (Biz)  [BR-08]", result.team_velocity_biz),
        ("QC − Dev  [BR-09]", result.qc_minus_dev),
    ]
    for kpi, val in kpis:
        r += 1
        ws.cell(r, 1, kpi)
        ws.cell(r, 2, val)

    if result.warnings:
        r += 2
        ws.cell(r, 1, "Warnings / Errors").font = bf
        _, _, wfill = _styles()
        for w in result.warnings:
            r += 1
            cell = ws.cell(r, 1, w)
            cell.fill = wfill
    _autowidth(ws)

    # ── Sheet 2: Resources ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resources")
    _hrow(ws2, 1, ["Resource", "Type", "Leave Days", "OT Days", "FTE noOT", "Full V", "V", "V OT"])
    for i, rr in enumerate(result.resource_results, 2):
        ws2.cell(i, 1, rr.display_name)
        ws2.cell(i, 2, rr.type.value)
        ws2.cell(i, 3, rr.leave_days)
        ws2.cell(i, 4, rr.ot_days)
        ws2.cell(i, 5, rr.fte_no_ot)
        ws2.cell(i, 6, rr.full_v)
        ws2.cell(i, 7, rr.v)
        ws2.cell(i, 8, rr.v_ot)
    _autowidth(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CSV export ───────────────────────────────────────────────────────────────

def result_to_csv(result: CalculationResult) -> bytes:
    """Export resource breakdown as CSV bytes."""
    lines = ["Resource,Type,Leave Days,OT Days,FTE noOT,Full V,V,V OT"]
    for rr in result.resource_results:
        lines.append(
            f"{rr.display_name},{rr.type.value},"
            f"{rr.leave_days},{rr.ot_days},"
            f"{rr.fte_no_ot},{rr.full_v},{rr.v},{rr.v_ot}"
        )
    return "\n".join(lines).encode("utf-8")
