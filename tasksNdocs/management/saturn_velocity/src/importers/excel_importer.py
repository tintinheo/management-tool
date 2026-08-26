"""
Excel importer — detects schema version per sheet and maps to domain models.

Schema detection strategy (per Technical Proposal §8):
  - s192: headers contain "velocity per day", "leave days", "ot in hours", "type"
  - s158: headers contain "resource dev" or "resource qc"
  - unknown: everything else

Import is non-destructive: unrecognised sheets are skipped with a warning, never silently dropped.
"""
import io
import uuid
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from ..domain.models import (
    LeaveEvent, LeaveStatus, ResourceType, Scenario, ScenarioResource, Sprint, Ticket,
)

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ── Schema signatures ────────────────────────────────────────────────────────

_S192_RES_ANCHORS = {"velocity per day", "leave days", "ot in hours", "type"}
_S158_RES_ANCHORS = {"resource dev", "resource qc"}

# ── Low-level helpers ────────────────────────────────────────────────────────

def _cell(ws, row: int, col: int) -> Any:
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None


def _to_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
    except ValueError:
        pass
    try:
        return datetime.strptime(str(val).strip(), "%d/%m/%Y").date()
    except ValueError:
        pass
    return None


def _to_float(val: Any, default: float = 0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _to_int(val: Any, default: int = 0) -> int:
    f = _to_float(val, float(default))
    return int(round(f))


def _row_tokens(ws, row: int, max_col: int = 25) -> set:
    tokens = set()
    for c in range(1, max_col + 1):
        v = _cell(ws, row, c)
        if v is not None:
            tokens.add(str(v).lower().strip())
    return tokens


# ── Schema detection ─────────────────────────────────────────────────────────

def detect_schema(ws) -> str:
    """Return 's192', 's158', or 'unknown'."""
    for r in range(1, 70):
        tokens = _row_tokens(ws, r)
        if len(tokens & _S192_RES_ANCHORS) >= 3:
            return "s192"
        if len(tokens & _S158_RES_ANCHORS) >= 1:
            return "s158"
    return "unknown"


# ── Sprint metadata extraction ────────────────────────────────────────────────

def _find_meta_anchor_row(ws) -> Optional[int]:
    """
    Locate the row that begins the sprint metadata block.
    Strategy: find a cell containing 'start' and 'date' (or just 'start') in column 1,
    between rows 30 and 65.
    """
    for r in range(30, 66):
        v = str(_cell(ws, r, 1) or "").lower()
        if ("start" in v and "date" in v) or v.strip() == "start":
            return r
    # Fallback: find any row with a date-like value in column 2 around row 36
    for r in range(34, 45):
        v = _cell(ws, r, 2)
        if isinstance(v, (date, datetime)):
            return r
    return None


def import_sprint(ws, sheet_name: str) -> Tuple[Optional[Sprint], List[str]]:
    warnings: List[str] = []
    meta_row = _find_meta_anchor_row(ws)
    if meta_row is None:
        warnings.append(f"[{sheet_name}] Could not locate sprint metadata block.")
        return None, warnings

    start_date = _to_date(_cell(ws, meta_row, 2))
    end_date = _to_date(_cell(ws, meta_row + 1, 2))
    dev_end = _to_date(_cell(ws, meta_row + 2, 2)) or end_date
    public_h = _to_int(_cell(ws, meta_row + 3, 2), 0)
    buffer_v = _to_float(_cell(ws, meta_row + 4, 2), 0.1)
    backup_v = _to_float(_cell(ws, meta_row + 5, 2), 1.0)

    if start_date is None or end_date is None:
        warnings.append(f"[{sheet_name}] Could not parse start/end dates from meta block (row {meta_row}).")
        return None, warnings

    if buffer_v > 1.0:
        buffer_v = buffer_v / 100.0  # normalise if stored as percentage integer
    if dev_end is None:
        dev_end = end_date

    sprint = Sprint(
        sprint_id=str(uuid.uuid4())[:8],
        name=sheet_name,
        start_date=start_date,
        end_date=end_date,
        development_end_date=dev_end,
        public_holidays=public_h,
        buffer=buffer_v,
        backup=backup_v,
    )
    return sprint, warnings


# ── Resource extraction (s192) ───────────────────────────────────────────────

def _build_col_map(ws, header_row: int) -> Dict[str, int]:
    """Map canonical field names to 1-based column indices from a header row."""
    col_map: Dict[str, int] = {}
    for c in range(1, 25):
        raw = str(_cell(ws, header_row, c) or "").lower().strip()
        if not raw:
            continue
        if "velocity" in raw and "day" in raw:
            col_map["velocity"] = c
        elif "leave" in raw and "day" in raw:
            col_map["leave_days"] = c
        elif "ot" in raw and "hour" in raw:
            col_map["ot_hours"] = c
        elif "ot" in raw and "day" in raw:
            col_map["ot_days"] = c
        elif raw == "v%":
            col_map["v_percent"] = c
        elif raw == "others":
            col_map["others"] = c
        elif raw == "type":
            col_map["type"] = c
        elif raw in ("resource", "name", "member"):
            col_map["name"] = c
    if "name" not in col_map:
        col_map["name"] = 1  # fallback: first column
    return col_map


def _find_resource_header_row(ws) -> Optional[int]:
    for r in range(40, 70):
        tokens = _row_tokens(ws, r)
        if len(tokens & {"velocity per day", "leave days"}) >= 1:
            return r
    return None


def import_resources_s192(
    ws, scenario_id: str, sheet_name: str
) -> Tuple[List[ScenarioResource], List[str]]:
    warnings: List[str] = []
    resources: List[ScenarioResource] = []

    header_row = _find_resource_header_row(ws)
    if header_row is None:
        warnings.append(f"[{sheet_name}] Could not find resource header row.")
        return resources, warnings

    col_map = _build_col_map(ws, header_row)
    if "velocity" not in col_map:
        warnings.append(f"[{sheet_name}] 'velocity per day' column not found; skipping resources.")
        return resources, warnings

    for r in range(header_row + 1, header_row + 20):
        name_val = _cell(ws, r, col_map["name"])
        if not name_val:
            continue
        name = str(name_val).strip()
        if not name or name.lower() in ("total", "sum", "grand total", ""):
            continue

        def _f(key: str, default: float = 0.0) -> float:
            if key in col_map:
                return _to_float(_cell(ws, r, col_map[key]), default)
            return default

        raw_type = str(_cell(ws, r, col_map.get("type", 0)) or "Dev").strip()
        try:
            rtype = ResourceType(raw_type)
        except ValueError:
            rtype = ResourceType.DEV
            warnings.append(
                f"[{sheet_name}] Resource '{name}': unknown type '{raw_type}', defaulting to Dev."
            )

        v_pct = _f("v_percent", 1.0)
        if v_pct > 1.5:
            v_pct = v_pct / 100.0  # stored as integer percentage

        resources.append(
            ScenarioResource(
                scenario_id=scenario_id,
                resource_id=str(uuid.uuid4())[:8],
                display_name=name,
                velocity=_f("velocity", 1.0),
                leave_days=_f("leave_days", 0.0),
                ot_hours=_f("ot_hours", 0.0),
                ot_days=_f("ot_days", 0.0),
                v_percent=v_pct,
                others=_f("others", 0.0),
                type=rtype,
            )
        )

    return resources, warnings


# ── Public API ───────────────────────────────────────────────────────────────

def import_workbook(file_bytes: bytes) -> Dict:
    """
    Import an XLSX workbook.

    Returns
    -------
    {
        "sprints":       [Sprint, ...],
        "scenarios":     [Scenario, ...],
        "import_report": {
            "sheets":   [{"sheet": str, "schema": str, "status": str, "warnings": [...]}],
            "warnings": [str],
            "errors":   [str],
        }
    }
    """
    if not _HAS_OPENPYXL:
        return {
            "sprints": [], "scenarios": [],
            "import_report": {
                "sheets": [],
                "warnings": ["openpyxl is not installed."],
                "errors": ["Missing dependency: openpyxl"],
            },
        }

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    sprints: List[Sprint] = []
    scenarios: List[Scenario] = []
    sheet_reports = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        schema = detect_schema(ws)
        entry: Dict[str, Any] = {
            "sheet": sheet_name, "schema": schema,
            "status": "ok", "warnings": [],
        }

        sprint, sw = import_sprint(ws, sheet_name)
        entry["warnings"].extend(sw)

        if sprint is None:
            entry["status"] = "skipped"
            entry["reason"] = "Could not extract sprint metadata"
            sheet_reports.append(entry)
            continue

        sprints.append(sprint)
        scenario_id = str(uuid.uuid4())[:8]
        scenario = Scenario(
            scenario_id=scenario_id,
            sprint_id=sprint.sprint_id,
            name=sheet_name,
        )

        if schema == "s192":
            res, rw = import_resources_s192(ws, scenario_id, sheet_name)
            entry["warnings"].extend(rw)
            scenario.resources = res
            entry["resources_imported"] = len(res)
        elif schema == "s158":
            entry["warnings"].append(
                f"[{sheet_name}] s158 schema detected — resource detail not imported "
                "(legacy adapter not yet available). Sprint metadata imported."
            )
        else:
            entry["warnings"].append(
                f"[{sheet_name}] Unknown schema — resource rows not imported."
            )

        scenarios.append(scenario)
        sheet_reports.append(entry)

    all_warnings = [w for s in sheet_reports for w in s.get("warnings", [])]
    all_errors = [w for w in all_warnings if w.startswith("ERROR")]

    return {
        "sprints": sprints,
        "scenarios": scenarios,
        "import_report": {
            "sheets": sheet_reports,
            "warnings": all_warnings,
            "errors": all_errors,
        },
    }
